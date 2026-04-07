from __future__ import annotations

from io import BytesIO
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ExcelValue = str | int | float | bool | None


def build_workbook(
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[ExcelValue]],
) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sanitize_sheet_name(sheet_name)
    worksheet.append(list(headers))

    for row in rows:
        worksheet.append(list(row))

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = max(len(str(cell.value or "")) for cell in column_cells) + 3
        worksheet.column_dimensions[column_letter].width = min(max(width, 14), 48)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def sanitize_sheet_name(value: str) -> str:
    sanitized = "".join(char for char in value if char not in r"[]:*?/\\").strip()
    return (sanitized or "Relatorio")[:31]
