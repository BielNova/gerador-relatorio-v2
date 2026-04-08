from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.ai import AssistantIntent
from backend.app.encodings import decode_mixed_text
from backend.app.main import app, assistant_service


client = TestClient(app)


def test_decode_mixed_text_handles_cp1252_and_utf8() -> None:
    assert decode_mixed_text(b"ESCRIT\xd3RIO") == "ESCRITÓRIO"
    assert decode_mixed_text("CAMINHÃO".encode("utf-8")) == "CAMINHÃO"


def test_companies_endpoint_lists_allowed_schemas() -> None:
    response = client.get("/api/companies")
    assert response.status_code == 200

    payload = response.json()
    assert [item["id"] for item in payload] == ["emp0001", "emp0002", "emp0003", "emp0004"]
    assert payload[0]["hasData"] is True
    assert all(item["hasData"] for item in payload)


def test_invalid_company_is_rejected() -> None:
    response = client.get("/api/reports/products-finished", params={"company": "public"})
    assert response.status_code == 400
    assert "Empresa inválida" in response.json()["detail"]


def test_company_overview_separates_business_areas() -> None:
    response = client.get("/api/dashboard/overview", params={"company": "emp0001"})
    assert response.status_code == 200

    payload = response.json()
    areas = {area["id"]: area for area in payload["areas"]}

    assert payload["summary"]["areaCount"] == 11
    assert payload["summary"]["activeAreaCount"] >= 8
    assert areas["comercial"]["label"] == "Comercial"
    assert areas["comercial"]["totalRows"] == 4076536
    assert areas["financeiro"]["totalRows"] == 1201127
    assert areas["financeiro"]["entryPath"] == "/finance"
    assert areas["producao"]["totalRows"] == 3289232
    assert areas["fiscal"]["entryPath"] == "/ncm-tax-rates"
    assert areas["estoque"]["entryPath"] == "/products"


def test_company_overview_rejects_invalid_company() -> None:
    response = client.get("/api/dashboard/overview", params={"company": "public"})

    assert response.status_code == 400


def test_products_finished_report_matches_current_database_snapshot() -> None:
    response = client.get("/api/reports/products-finished", params={"company": "emp0001"})
    assert response.status_code == 200

    payload = response.json()
    assert payload["company"] == "emp0001"
    assert payload["summary"] == {
        "totalRows": 3390,
        "missingNcmRows": 293,
        "groupCount": 21,
    }
    assert payload["rows"][0] == {
        "code": "01010004",
        "description": "AMACIANTE ABRAÇO",
        "group": "AMACIANTES",
        "ncm": "38099190",
        "icmsRate": 20.5,
        "ipiRate": 0.0,
        "pisRate": 1.65,
        "cofinsRate": 7.6,
    }


def test_products_finished_report_returns_empty_state_for_company_without_data() -> None:
    response = client.get("/api/reports/products-finished", params={"company": "emp0002"})
    assert response.status_code == 200

    payload = response.json()
    assert payload["summary"] == {
        "totalRows": 0,
        "missingNcmRows": 0,
        "groupCount": 0,
    }
    assert payload["rows"] == []


def test_ncm_tax_rates_report_matches_current_database_snapshot() -> None:
    response = client.get("/api/reports/ncm-tax-rates", params={"company": "emp0001"})
    assert response.status_code == 200

    payload = response.json()
    assert payload["company"] == "emp0001"
    assert payload["summary"] == {
        "totalRows": 1197,
        "distinctNcms": 684,
        "duplicateNcmRows": 513,
    }
    assert payload["rows"][0] == {
        "ncm": "00000000",
        "icmsRate": 0.0,
        "ipiRate": 0.0,
        "pisRate": 1.65,
        "cofinsRate": 7.6,
    }


def test_ncm_tax_rates_report_returns_empty_state_for_company_without_data() -> None:
    response = client.get("/api/reports/ncm-tax-rates", params={"company": "emp0002"})
    assert response.status_code == 200

    payload = response.json()
    assert payload["summary"] == {
        "totalRows": 0,
        "distinctNcms": 0,
        "duplicateNcmRows": 0,
    }
    assert payload["rows"] == []


def test_ncm_tax_rates_export_downloads_xlsx() -> None:
    response = client.get(
        "/api/reports/ncm-tax-rates/export.xlsx",
        params={"company": "emp0001", "search": "38099190"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="arquimedes-ncm-aliquotas-emp0001.xlsx"' in response.headers[
        "content-disposition"
    ]
    assert response.content.startswith(b"PK")


def test_products_finished_export_includes_tax_rates_with_products() -> None:
    response = client.get(
        "/api/reports/products-finished/export.xlsx",
        params={"company": "emp0001", "search": "01010004"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    first_row = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]

    assert headers == [
        "Codigo Produto Acabado",
        "Descricao Produto Acabado",
        "Grupo",
        "NCM",
        "Aliquota ICMS",
        "Aliquota IPI",
        "Aliquota PIS",
        "Aliquota COFINS",
    ]
    assert first_row[0] == "01010004"
    assert str(first_row[1]).startswith("AMACIANTE ABRA")
    assert first_row[2:] == ["AMACIANTES", "38099190", 20.5, 0, 1.65, 7.6]


def test_fiscal_dashboard_matches_current_database_snapshot() -> None:
    response = client.get("/api/dashboard/fiscal", params={"company": "emp0001"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"] == {
        "totalProducts": 3390,
        "missingNcmProducts": 293,
        "distinctNcms": 684,
        "ncmVariationCount": 286,
        "duplicateNcmRows": 513,
        "productsWithAnyZeroRate": 971,
        "zeroIcmsProducts": 308,
        "zeroIpiProducts": 677,
        "zeroPisProducts": 293,
        "zeroCofinsProducts": 293,
    }
    assert payload["groupIssues"][0]["group"] == "DESCONTINUALIZADOS"


def test_fiscal_dashboard_rejects_invalid_company() -> None:
    response = client.get("/api/dashboard/fiscal", params={"company": "public"})

    assert response.status_code == 400


def test_finance_receivables_report_matches_current_database_snapshot() -> None:
    response = client.get("/api/finance/receivables", params={"company": "emp0001"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["company"] == "emp0001"
    assert payload["summary"]["totalOpenRows"] == 3871
    assert payload["summary"]["overdueRows"] == 2064
    assert payload["summary"]["filteredRows"] == 3449
    assert payload["rows"][0]["boletoCode"] == "0000000191"
    assert payload["rows"][0]["personName"] == "CONSUMIDOR"
    assert payload["rows"][0]["statusLabel"] == "Em aberto"
    assert payload["topDebtors"][0]["personName"] == "C.M.C. PRODUTOS QUIMICOS LTDA"


def test_finance_dashboard_matches_current_database_snapshot() -> None:
    response = client.get(
        "/api/finance/dashboard",
        params={"company": "emp0001", "referenceDate": "2026-04-07"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["company"] == "emp0001"
    assert payload["referenceDate"] == "2026-04-07"
    assert payload["period"] == {
        "mode": "month",
        "label": "Este mes",
        "startDate": "2026-04-01",
        "endDate": "2026-04-07",
        "cashFlow": {
            "label": "Este mes",
            "startDate": "2026-04-01",
            "endDate": "2026-04-07",
            "inflow": 306737.13,
            "outflow": 419742.66,
            "net": -113005.53,
        },
        "payablesDue": {"rows": 132, "amount": 437857.11},
        "receivablesDue": {"rows": 249, "amount": 315170.47},
        "received": {"rows": 0, "amount": 0.0},
    }
    assert payload["cash"] == {
        "sourceDate": "2026-03-25",
        "currentCash": -346343.59,
        "consolidatedBalance": -691079.33,
        "availableCash": -3608076.48,
        "committedCash": 3261732.89,
    }
    assert payload["payables"]["open"] == {"rows": 2202, "amount": 11051306.87}
    assert payload["payables"]["overdue"] == {"rows": 191, "amount": 821165.79}
    assert payload["payables"]["dueToday"] == {"rows": 57, "amount": 130183.05}
    assert payload["receivables"]["open"] == {"rows": 3871, "amount": 5179236.06}
    assert payload["receivables"]["overdue"] == {"rows": 2064, "amount": 2276509.14}
    assert payload["dre"]["year"] == 2026
    assert payload["dre"]["month"] == 2
    assert payload["dre"]["isFallbackMonth"] is True
    assert payload["dre"]["revenueTotal"] == 1929985.75
    assert payload["dre"]["netProfit"] == -162416.36
    assert payload["indicators"]["profitabilityPercent"] == -8.42
    assert payload["topDebtors"][0]["personName"] == "C.M.C. PRODUTOS QUIMICOS LTDA"
    assert payload["alerts"][0]["title"] == "Contas vencidas"
    assert payload["audit"]["cash"]["source"] == "FNCBMOV + FNCBLANC + FNCTBCO"
    assert payload["audit"]["cash"]["amount"] == -346343.59
    assert payload["audit"]["payables"]["rawOpen"] == {
        "rows": 2202,
        "amount": 11051306.87,
        "startDate": "2025-12-04",
        "endDate": "4202-07-13",
    }
    assert payload["audit"]["payables"]["currentYearOpen"]["rows"] == 1109
    assert payload["audit"]["payables"]["currentYearOpen"]["amount"] == 5523971.8
    assert payload["audit"]["payables"]["futureAnomalies2030Plus"]["rows"] == 568
    assert payload["audit"]["payables"]["futureAnomalies2030Plus"]["amount"] == 838970.69
    assert payload["audit"]["payables"]["missingDueDate"] == {
        "rows": 8,
        "amount": 208.76,
        "startDate": None,
        "endDate": None,
    }
    assert payload["audit"]["receivables"]["rawOpen"]["rows"] == 3871
    assert payload["audit"]["receivables"]["rawOpen"]["amount"] == 5179236.06
    assert payload["audit"]["receivables"]["expected30Days"]["amount"] == 1985076.73
    assert payload["audit"]["dre"] == {
        "dreReferenceMonth": "2026-02",
        "latestInvoiceMonthFromCVFATURA": "2026-03",
        "isFallbackMonth": True,
        "isStaleComparedToInvoices": True,
    }
    assert any(alert["title"] == "DRE desatualizada na base" for alert in payload["alerts"])


def test_finance_dashboard_period_modes() -> None:
    response = client.get(
        "/api/finance/dashboard",
        params={"company": "emp0001", "referenceDate": "2026-04-07", "period": "year"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["period"]["mode"] == "year"
    assert payload["period"]["label"] == "Este ano"
    assert payload["period"]["startDate"] == "2026-01-01"
    assert payload["period"]["endDate"] == "2026-04-07"
    assert payload["period"]["cashFlow"]["net"] == -818098.72
    assert payload["period"]["received"] == {"rows": 4357, "amount": 5665248.78}


def test_finance_dashboard_rejects_invalid_period() -> None:
    response = client.get(
        "/api/finance/dashboard",
        params={"company": "emp0001", "period": "semester"},
    )

    assert response.status_code == 400
    assert "Periodo financeiro invalido" in response.json()["detail"]


def test_finance_dashboard_rejects_invalid_company() -> None:
    response = client.get("/api/finance/dashboard", params={"company": "public"})

    assert response.status_code == 400


def test_finance_receivables_report_filters_overdue_and_customer() -> None:
    response = client.get(
        "/api/finance/receivables",
        params={
            "company": "emp0001",
            "onlyOverdue": "true",
            "search": "C.M.C. PRODUTOS",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["filters"]["onlyOverdue"] is True
    assert payload["summary"]["filteredRows"] == 279
    assert payload["rows"][0]["personName"] == "C.M.C. PRODUTOS QUIMICOS LTDA"
    assert all(row["daysOverdue"] > 0 for row in payload["rows"])


def test_finance_receivables_report_rejects_invalid_company() -> None:
    response = client.get("/api/finance/receivables", params={"company": "public"})

    assert response.status_code == 400


def test_finance_receivables_export_downloads_xlsx() -> None:
    response = client.get(
        "/api/finance/receivables/export.xlsx",
        params={"company": "emp0001", "onlyOverdue": "true", "search": "C.M.C. PRODUTOS"},
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    first_row = [cell.value for cell in next(worksheet.iter_rows(min_row=2, max_row=2))]

    assert headers[:6] == [
        "Boleto",
        "Titulo",
        "Contrato",
        "Parcela",
        "Codigo Pessoa",
        "Cliente",
    ]
    assert first_row[5] == "C.M.C. PRODUTOS QUIMICOS LTDA"
    assert 'filename="arquimedes-contas-a-receber-emp0001.xlsx"' in response.headers[
        "content-disposition"
    ]


def test_report_assistant_runs_allowed_intent_with_mocked_classifier() -> None:
    original_classifier = assistant_service.classifier
    assistant_service.classifier = StaticClassifier(
        AssistantIntent(intent="products_by_ncm", params={"ncm": "38099190"})
    )
    try:
        response = client.post(
            "/api/ai/report-assistant",
            json={"company": "emp0001", "question": "produtos do NCM 38099190"},
        )
    finally:
        assistant_service.classifier = original_classifier

    assert response.status_code == 200
    payload = response.json()
    assert payload["intent"] == "products_by_ncm"
    assert payload["totalRows"] > 0
    assert payload["columns"][2:] == ["Grupo", "NCM", "ICMS", "IPI", "PIS", "COFINS"]
    assert payload["columns"][0].startswith("C")
    assert payload["columns"][1].startswith("Descri")
    assert "export.xlsx" in payload["exportUrl"]


def test_report_assistant_runs_finance_intent_with_mocked_classifier() -> None:
    original_classifier = assistant_service.classifier
    assistant_service.classifier = StaticClassifier(
        AssistantIntent(intent="finance_receivables_by_customer", params={"customer": "C.M.C."})
    )
    try:
        response = client.post(
            "/api/ai/report-assistant",
            json={"company": "emp0001", "question": "boletos vencidos do cliente C.M.C."},
        )
    finally:
        assistant_service.classifier = original_classifier

    assert response.status_code == 200
    payload = response.json()
    assert payload["intent"] == "finance_receivables_by_customer"
    assert payload["totalRows"] > 0
    assert payload["columns"] == [
        "Boleto",
        "Titulo",
        "Contrato",
        "Parcela",
        "Cliente",
        "Vencimento",
        "Valor",
        "Dias vencidos",
        "Status",
    ]
    assert "/api/finance/receivables/export.xlsx" in payload["exportUrl"]


def test_report_assistant_rejects_unsupported_intent() -> None:
    original_classifier = assistant_service.classifier
    assistant_service.classifier = StaticClassifier(AssistantIntent(intent="unsupported", params={}))
    try:
        response = client.post(
            "/api/ai/report-assistant",
            json={"company": "emp0001", "question": "faÃ§a qualquer SQL"},
        )
    finally:
        assistant_service.classifier = original_classifier

    assert response.status_code == 422


def test_report_assistant_rejects_invalid_company_before_ai() -> None:
    response = client.post(
        "/api/ai/report-assistant",
        json={"company": "public", "question": "produtos sem NCM"},
    )

    assert response.status_code == 400


class StaticClassifier:
    def __init__(self, intent: AssistantIntent) -> None:
        self.intent = intent

    def classify(self, question: str) -> AssistantIntent:
        return self.intent
